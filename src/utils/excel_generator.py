from openpyxl import Workbook
from openpyxl.drawing.image import Image
# from openpyxl.styles import Alignment


class ExcelGenerator:
    def __init__(self, config):
        self.config = config
        self.workbook = Workbook()

    def generate(self, data):
        self._create_summary_sheet(data)
        self._create_category_sheets(data)
        self._add_vba_macros()
        self._save_workbook()

    def _create_summary_sheet(self, data):
        sheet = self.workbook.active
        sheet.title = "Summary"
        headers = ["URL", "Category", "Thumbnail", "Linked URLs", "Screenshot"]
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col, value=header)
        for row, item in enumerate(data, start=2):
            sheet.cell(row=row, column=1, value=item["url"])
            sheet.cell(row=row, column=2, value=item["category"])
            self._add_thumbnail(sheet, row, 3, item["thumbnail"])
            sheet.cell(row=row, column=4, value=", ".join(item["linked_urls"]))
            self._add_screenshot_link(sheet, row, 5, item["screenshot_path"])

    def _create_category_sheets(self, data):
        categories = set(item["category"] for item in data)
        for category in categories:
            sheet = self.workbook.create_sheet(title=category)
            # カテゴリシートの内容を作成

    def _add_thumbnail(self, sheet, row, col, thumbnail_path):
        img = Image(thumbnail_path)
        img.width = 100
        img.height = 100
        sheet.add_image(img, f"{chr(64+col)}{row}")

    def _add_screenshot_link(self, sheet, row, col, screenshot_path):
        cell = sheet.cell(row=row, column=col)
        cell.value = "View"
        cell.hyperlink = screenshot_path
        cell.style = "Hyperlink"

    def _add_vba_macros(self):
        # VBAマクロを追加するロジックをここに実装
        pass

    def _save_workbook(self):
        filename = f"{self.config.get('output', 'excel_prefix')}_{self.config.get_timestamp()}.xlsm"
        self.workbook.save(filename)
